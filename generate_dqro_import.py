"""Generate a CDGC DQRO bulk-import file from the template, keyed on each column's
Reference ID (primary data element path). --limit-tables for a test file."""
import argparse, glob, json, os, re
import openpyxl

TEMPLATE = "templates/CDGC_DQRO_TEMPLATE.xlsx"
CACHE = ".scan_cache"

# Mirrors ai_governance_mcp._select_key_columns / create_generic_dq_rules so the
# bulk file matches exactly what the UI demo creates ("potential columns" only).
_ID_PATTERNS   = ("_ID", "ID_", "_KEY", "KEY_", "PK_", "_PK", "CODE", "_NO", "NO_", "NBR", "_NUM", "NUM_")
_DATE_PATTERNS = ("_DATE", "DATE_", "_DT", "DT_", "_TIME", "TIME_", "CREATED", "MODIFIED", "UPDATED")
_TYPE_PATTERNS = ("_TYPE", "TYPE_", "_STATUS", "STATUS_", "_NAME", "NAME_", "_DESC", "DESC_", "_CAT", "CAT_")


def select_key_columns(columns, max_cols=7):
    """Curated subset of columns for DQ, priority-bucketed. Mirrors the UI demo."""
    buckets = {1: [], 2: [], 3: [], 4: [], 5: []}
    for col in columns:
        dt = (col.get("data_type") or "unknown").lower()
        name = col.get("name") or col.get("column_name")
        if dt == "unknown" or not name:
            continue
        cn = name.upper()
        if any(p in cn for p in _ID_PATTERNS):
            buckets[1].append(col)
        elif any(x in dt for x in ("timestamp", "date", "time")) or any(p in cn for p in _DATE_PATTERNS):
            buckets[2].append(col)
        elif any(p in cn for p in _TYPE_PATTERNS):
            buckets[3].append(col)
        elif any(x in dt for x in ("number", "numeric", "int", "decimal", "float", "double")):
            buckets[4].append(col)
        else:
            buckets[5].append(col)
    selected = []
    for pr in sorted(buckets):
        for col in buckets[pr]:
            if len(selected) >= max_cols:
                return selected
            selected.append(col)
    return selected


def dims_for(dtype: str, name: str):
    """Completeness + one dimension, per create_generic_dq_rules auto-selection."""
    d = ["Completeness"]; u = (dtype or "").upper(); n = (name or "").upper()
    if any(k in u for k in ("NUMBER", "NUMERIC", "INT", "DECIMAL", "FLOAT", "DOUBLE")):
        d.append("Uniqueness" if any(p in n for p in _ID_PATTERNS) else "Validity")
    elif "DATE" in u or "TIMESTAMP" in u or "TIME" in u:
        d.append("Timeliness")
    else:
        d.append("Validity")
    return d

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--limit-tables", type=int, default=0, help="0 = all cached tables")
    ap.add_argument("--out", default="templates/CDGC_DQRO_FILLED_TEST.xlsx")
    ap.add_argument("--rule-map", default="rule_map.json",
                    help="JSON {dimension: rule_spec_id} for Technical Rule Reference")
    ap.add_argument("--max-cols", type=int, default=7,
                    help="Max curated columns per table (UI demo default 7)")
    ap.add_argument("--origin-filter", default=None,
                    help="Only include cache tables whose external_id contains this "
                         "substring (e.g. GOVERNANCE_SCALE_TEST) to exclude stale cache")
    ap.add_argument("--operation", default="Create", choices=["Create", "Delete", "Update"],
                    help="CDGC bulk-import Operation column value (default Create)")
    args = ap.parse_args()

    # Technical Rule Reference = FRS document id of the CDQ rule spec (mandatory).
    with open(args.rule_map) as f:
        rule_map = json.load(f)
    missing = [d for d in ("Completeness", "Validity", "Uniqueness",
                           "Timeliness", "Accuracy", "Consistency")
               if not rule_map.get(d)]
    if missing:
        raise SystemExit(f"rule_map missing rule spec ids for: {missing}. "
                         f"Run probe_rule_specs.py first.")

    wb = openpyxl.load_workbook(TEMPLATE)
    ws = wb["Data Quality Rule Occurrence"]
    hdr = [c.value for c in ws[1]]
    idx = {h: i for i, h in enumerate(hdr)}

    files = sorted(glob.glob(os.path.join(CACHE, "*.json")))
    if args.limit_tables:
        files = files[:args.limit_tables]
    n = 0
    for cf in files:
        d = json.load(open(cf))
        ext = d.get("external_id", "")
        if "~" not in ext:
            continue
        if args.origin_filter and args.origin_filter not in ext:
            continue
        base = ext.split("~")[0]          # origin://DB/SCHEMA/TABLE
        table = d.get("name", "")
        # UI-demo parity: only the curated "potential" columns (max 7/table).
        key_cols = select_key_columns(
            [c for c in d.get("columns", []) if not (c.get("name") or "").startswith("SYS_")],
            max_cols=args.max_cols,
        )
        for col in key_cols:
            cname = col.get("name"); dtype = col.get("data_type", "")
            if not cname:
                continue
            ref = f"{base}/{cname}~com.infa.odin.models.relational.Column"
            for dim in dims_for(dtype, cname):
                row = [None] * len(hdr)
                row[idx["Name"]] = f"DQ_{table}_{cname}_{dim}"[:120]
                row[idx["Criticality"]] = "High"
                row[idx["Dimension"]] = dim
                row[idx["Lifecycle"]] = "Published"
                row[idx["Measuring Method"]] = "InformaticaCloudDataQuality"
                row[idx["Technical Description"]] = f"{dim} check on {table}.{cname}"
                row[idx["Target"]] = 95
                row[idx["Threshold"]] = 80
                row[idx["Primary Data Element"]] = ref
                # Technical Rule Reference = FRS id of the CDQ rule spec for this dimension.
                row[idx["Technical Rule Reference"]] = rule_map[dim]
                if "Input Port Name" in idx:
                    row[idx["Input Port Name"]] = "Input"
                if "Output Port Name" in idx:
                    row[idx["Output Port Name"]] = "PrimaryRuleSet"
                row[idx["Operation"]] = args.operation
                ws.append(row); n += 1
    wb.save(args.out)
    print(f"wrote {n} DQRO rows across {len(files)} tables -> {args.out}")
    # show a sample row
    if n:
        r = list(ws.iter_rows(values_only=True))[1]
        for h, v in zip(hdr, r):
            if v is not None:
                print(f"   {h}: {str(v)[:90]}")

if __name__ == "__main__":
    main()
